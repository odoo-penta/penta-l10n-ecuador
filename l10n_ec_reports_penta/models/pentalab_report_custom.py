# -*- coding: utf-8 -*-
from datetime import datetime
import io
import openpyxl
from odoo import models, fields, api
import re
import logging
import json
from openpyxl.styles import Alignment
from openpyxl.styles import Font
_logger = logging.getLogger(__name__)

class PentalabReportCustom(models.Model):
    _name = 'pentalab.report.custom'
    _description = 'Reportes modificados Pentalab'

    def generar_reporte_modificado(self, company_id, report_id, options):

        company = self.env['res.company'].browse(company_id)
        report = self.env['account.report'].with_company(company).with_context(lang='es_EC').browse(report_id)

        # 1) Exportamos y guardamos el archivo original
        xlsx_data_mensual = report.export_to_xlsx(options)
        workbook_mensual = openpyxl.load_workbook(io.BytesIO(xlsx_data_mensual['file_content']))
        file = 'reporte_modificado.xlsx'
        workbook_mensual.save(file)

        # 2) Reabrimos el archivo para modificarlo
        wb = openpyxl.load_workbook(file)
        ws = wb.active  # Usamos la primera hoja

        # -----------------------------------------------
        # Buscar estilo base de la primera celda no vacía en la col. A (fila >= 4)
        # -----------------------------------------------
        estilo_base = None
        for row in range(4, ws.max_row + 1):
            cell_a = ws.cell(row=row, column=1)
            if cell_a.value:
                estilo_base = cell_a._style
                break  # Salimos al primer match

        # -----------------------------------------------
        # PRIMER FOR:
        # - Eliminar fila si B empieza con "total" (lstrip, lower)
        # - Eliminar fila si C es 0 o '0.00'
        # - Si A está vacía y B y C tienen valor, eliminar la fila si la fila siguiente tiene el mismo valor en C
        # - Separar texto B -> A si A está vacío y B empieza con número
        # -----------------------------------------------
        for row in range(ws.max_row, 3, -1):  # desde abajo hacia la fila 4
            cell_b = ws.cell(row=row, column=2)
            cell_a = ws.cell(row=row, column=1)
            cell_c = ws.cell(row=row, column=3)

            # -- 1) Eliminar fila si B comienza con "total"
            if cell_b.value and isinstance(cell_b.value, str) and cell_b.value.lstrip().lower().startswith("total"):
                ws.delete_rows(row, 1)
                continue

            # -- 2) Eliminar fila si C == 0 (numérico o string "0", "0.00", etc.)
            c_value = cell_c.value
            if self._is_zero_value(c_value):
                ws.delete_rows(row, 1)
                continue

            # -- 4) Si A está vacía, y B contiene un número al inicio, pasar esa parte a A
            if not cell_a.value and cell_b.value and isinstance(cell_b.value, str):
                original_b = cell_b.value.lstrip()
                first_space_idx = original_b.find(' ')
                if first_space_idx != -1:
                    first_part = original_b[:first_space_idx]
                    if re.match(r'^\d+', first_part):  # empieza con dígitos
                        new_a_value = first_part
                        new_b_value = original_b[first_space_idx:].lstrip()

                        cell_a.value = new_a_value
                        cell_b.value = new_b_value
                        # Aplicar estilo base a la celda A
                        if estilo_base:
                            cell_a._style = estilo_base
                            

        # -----------------------------------------------
        # SEGUNDO FOR: Eliminar filas con cuentas ocultas (hide_in_report)
        # -----------------------------------------------
        accounts_dict = {
            acc.code_store.strip(): acc
            for acc in self.env['account.account'].search(['|', ('code_store', '!=', False), ('code_store', '!=', '')])
        }
        for row in range(ws.max_row, 3, -1):
            cell_a = ws.cell(row=row, column=1)
            if cell_a.value and isinstance(cell_a.value, str):
                code = cell_a.value.strip()
                account = accounts_dict.get(code)
                if account and account.hide_in_report:
                    ws.delete_rows(row, 1)

        

        # -----------------------------------------------
        # TERCER FOR: Buscar "Ganancias del año actual"
        # - Reemplazar texto en B por "Resultados del año actual"
        # - Cambiar A por code_store de la cuenta equity_unaffected
        # - Eliminar 2 filas arriba/abajo si A está vacía
        # -----------------------------------------------

        for row in range(ws.max_row, 3, -1):
            cell_b = ws.cell(row=row, column=2)
            if cell_b.value and isinstance(cell_b.value, str) and "Ganancias del año actual" in cell_b.value:
                # Reemplazar texto en B
                new_text = cell_b.value.replace("Ganancias del año actual", "Resultados del año actual")
                cell_b.value = new_text
        
                # Guardar valores de la fila antes de eliminarla
                row_values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
                row_styles = [ws.cell(row=row, column=col)._style for col in range(1, ws.max_column + 1)]
        
                # Eliminar las filas siguientes (row+1 y row+2)
                ws.delete_rows(row+1, 1)
                ws.delete_rows(row+1, 1)
        
                # Eliminar la fila actual
                ws.delete_rows(row, 1)
        
                # Insertar la fila guardada en la nueva posición (3 filas abajo)
                new_row = row + 2
                ws.insert_rows(new_row)
        
                for col, value in enumerate(row_values, start=1):
                    ws.cell(row=new_row, column=col).value = value
                    ws.cell(row=new_row, column=col)._style = row_styles[col-1]  # Aplicar el estilo original
        # Variable para almacenar la fila donde se copiarán los valores
        nueva_fila = None
        
        # Buscar la celda en la columna B con "Pasivos + Capital"
        for row in range(ws.max_row, 3, -1):
            cell_b = ws.cell(row=row, column=2)
        
            if cell_b.value and isinstance(cell_b.value, str) and cell_b.value == "Pasivos + Capital":
                
                # Guardar valores originales de A, B y C
                valor_a = ws.cell(row=row, column=1).value
                valor_b = ws.cell(row=row, column=2).value
                valor_c = ws.cell(row=row, column=3).value
        
                # Guardar estilos originales de A, B y C
                estilo_a = ws.cell(row=row, column=1)._style
                estilo_b = ws.cell(row=row, column=2)._style
                estilo_c = ws.cell(row=row, column=3)._style
        
                # Modificar la celda B en la fila original
                ws.cell(row=row, column=2).value = "Pasivos + Patrimonio"
        
                # Copiar los valores y estilos 5 filas más abajo
                nueva_fila = row + 5
                ws.cell(nueva_fila, 1).value = valor_a  # Columna A
                ws.cell(nueva_fila, 2).value = "PASIVO +PATRIMONIO+RESULTADOS"  # Modificar columna B
                ws.cell(nueva_fila, 3).value = valor_c  # Columna C
        
                # Aplicar estilos copiados a las nuevas celdas
                ws.cell(nueva_fila, 1)._style = estilo_a
                ws.cell(nueva_fila, 2)._style = estilo_b
                ws.cell(nueva_fila, 3)._style = estilo_c
        
        # Variables para guardar los valores de C cuando A=3 y A=2
        valor_c_a3 = None
        valor_c_a2 = None
        
        # Buscar las filas donde A=3 y A=2
        for row in range(1, ws.max_row + 1):
            cell_a = ws.cell(row=row, column=1)
            cell_c = ws.cell(row=row, column=3)
            
            if cell_a.value == "3":
                valor_c_a3 = cell_c.value
            elif cell_a.value == "2":
                valor_c_a2 = cell_c.value
        
        # Si encontramos ambas celdas, sumamos sus valores
        if valor_c_a3 is not None and valor_c_a2 is not None and nueva_fila:
            suma_c = int(valor_c_a3) + int(valor_c_a2)
            ws.cell(nueva_fila-5, 3, suma_c)  # Guardamos la suma en la celda C de la fila copiada

        textos_a_eliminar = {
            "Ganancias sin asignar",
            "Ganancias no asignadas del año en curso",
            "Patrimonio Neto",
            "Ganancias acumuladas"
        }
        
        # Recorrer desde la última fila hasta la cuarta fila (para evitar encabezados)
        for row in range(ws.max_row, 3, -1):  
            cell_b = ws.cell(row=row, column=2)
            
            if cell_b.value and isinstance(cell_b.value, str) and cell_b.value.strip() in textos_a_eliminar:
                ws.delete_rows(row, 1)  # Elimina la fila completa

        # Antes de guardar el archivo, agregamos el encabezado con estilos
        company = self.env['res.company'].browse(company_id)
        
        # Insertar una fila extra después de la fila 2
        ws.insert_rows(1)
        
        # Estilo del encabezado
        header_font = Font(name="Arial", size=14, bold=True)
        
        # Fila 1 - Nombre de la empresa
        ws.cell(row=1, column=2).value = company.name
        ws.cell(row=1, column=2).font = header_font  # Aplicar estilo
        ws.cell(row=1, column=2).alignment = Alignment(horizontal="center")
        # Fila 2 - Nombre del reporte según report_id
        if options.get('report_id') == 23:
            report_name = "Balance General"
        elif options.get('report_id') == 24:
            report_name = "Estado de Resultados"
        else:
            report_name = "Reporte Financiero"
        ws.cell(row=2, column=2).value = report_name
        header_font = Font(name="Arial", size=12, bold=True)
        
        # Fila 3 - Fecha de generación del reporte
        fecha_generacion = datetime.today().strftime('%d-%m-%Y')
        
        
        ws.cell(row=2, column=3).font = header_font
        
        ws.cell(row=2, column=2).alignment = Alignment(horizontal="center")
        
        ws.cell(row=2, column=2).font = header_font  # Aplicar estilo
        header_font = Font(name="Arial", size=10)
        ws.cell(row=3, column=2).value = "Fecha de generación del reporte"
        ws.cell(row=3, column=3).value = fecha_generacion
        ws.cell(row=3, column=2).font = header_font   # Aplicar estilo
        ws.cell(row=3, column=3).font = header_font
        # Guardar cambios finales
        wb.save(file)


    # -----------------------------------------------
    # MÉTODOS DE APOYO
    # -----------------------------------------------
    def _is_zero_value(self, value):
        """Retorna True si el valor representa cero (0 o '0.00', etc.)"""
        if value is None:
            return False
        # Caso numérico
        if isinstance(value, (int, float)):
            return (value == 0)
        # Caso string: eliminar espacios y comprobar
        if isinstance(value, str):
            v = value.strip().replace(',', '')  # por si viene con coma
            # Admite "0", "0.00", "0.0" etc.
            try:
                return (float(v) == 0)
            except ValueError:
                return False
        return False

    def _compare_cell_values(self, val1, val2):
        """
        Compara dos valores de celda (numéricos o string) para ver si son "iguales".
        Permite strings '100.00' vs float 100.0, etc.
        """
        # Ambos None => iguales
        if val1 is None and val2 is None:
            return True
        # Uno None, otro no => distintos
        if val1 is None or val2 is None:
            return False

        # Convertir ambos a float si se puede
        try:
            f1 = float(str(val1).replace(',', ''))
            f2 = float(str(val2).replace(',', ''))
            return abs(f1 - f2) < 1e-9  # casi iguales
        except ValueError:
            # No se pudieron convertir a float => comparar como string
            return str(val1).strip() == str(val2).strip()
