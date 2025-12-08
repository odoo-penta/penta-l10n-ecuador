# -*- coding: utf-8 -*-
import base64
import io
from datetime import datetime, date
from io import BytesIO
import re
import openpyxl
import xlsxwriter

from datetime import timedelta

from odoo import models, fields

from odoo.tools.misc import format_date
import base64
from io import BytesIO
from openpyxl import load_workbook

# DATE_FORMAT = "%Y-%m-%d"
DATE_FORMAT = "%d/%m/%Y"


class PentalabReportWizard(models.TransientModel):
    _name = 'pentalab.report.wizard'
    _description = 'Wizard para Informe Pentalab'

    company_id = fields.Many2one(
        'res.company',
        string='Compañía',
        required=True,
        default=lambda self: self.env.company
    )
    date_from = fields.Date(string="Desde")
    date_to = fields.Date(string="Hasta", default=date.today())
    type = fields.Selection([
        ('out_invoice', 'Por Cobrar'),
        ('advance_account_customer_id', 'Por Cobrar y Por Pagar Anticipo de Clientes'),
        ('in_invoice', 'Por Pagar')
    ], string='Tipo', default='out_invoice', required=True)
    partner_id = fields.Many2one('res.partner', string="Socio")

    # Campos para almacenar el archivo generado
    file_data = fields.Binary("Archivo XLSX", readonly=True)
    file_name = fields.Char("Nombre del Archivo", readonly=True)

    def action_generate_report(self):

        company = self.env['res.company'].browse(self.company_id.id)
        report = self.env['account.report'].with_company(company).browse(9)

        hoy = datetime.now()
        if self.date_to:
            hoy = self.date_to
        date_from = datetime(2020, 1, 1)
        if self.date_from:
            date_from = self.date_from

        # Formatear las fechas
        date_now_formatted = hoy.strftime("%Y-%m-%d")
        date_from_formatted = date_from.strftime("%Y-%m-%d")

        # Obtener las opciones
        options = {
            'companies': [{'id': self.company_id.id}],  # Dinámico
            'single_company': [self.company_id.id],
            'date': {
                'date_from': date_from_formatted,
                'date_to': date_now_formatted,
            },
            'sections_source_id': 9,
            'export_mode': 'file',
            'all_entries': False,
            'journals': [],
            'show_account': True,
            'comparison': {
                'filter': 'no_comparison'
            },
            'unfold_all': True  # Propiedad para desplegar todas las cuentas
        }

        # Generar el archivo en binario usando export_to_xlsx para ambas opciones
        xlsx_data = report.export_to_xlsx(options)

        # Leer el archivo binario y cargarlo en openpyxl para la primera opción
        file_content1 = xlsx_data['file_content']
        excel_file1 = io.BytesIO(file_content1)
        workbook1 = openpyxl.load_workbook(excel_file1)

        # workbook1.save('balance_mensual.xlsx')

        sheet1 = workbook1.active  # Selecciona la primera hoja

        """Genera y descarga el informe XLSX con la lógica original (pagos, días vencidos, etc.)."""
        self.ensure_one()

        # 1) Construir el dominio
        domain = [
            ('company_id', '=', self.company_id.id),
            ('display_type', '=', 'payment_term'),
            ('amount_residual', '>', 0),
            ('move_id.move_type', '=', self.type),
            ('move_id.state', '=', 'posted'),
        ]

        if self.date_from:
            domain.append(('move_id.date', '>=', self.date_from))
        if self.date_to:
            domain.append(('move_id.date', '<=', self.date_to))
        if self.partner_id:
            domain.append(('partner_id', '=', self.partner_id.id))

        account_move_lines = self.env['account.move.line'].search(domain, order='name desc')

        # 2) Crear archivo en memoria con xlsxwriter
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output)
        sheet = workbook.add_worksheet('Reporte')
        bold = workbook.add_format({'bold': True})

        # Encabezado dinámico (fecha "Al XXXXX")
        start_col = 3
        if self.date_to:
            headersd = ['Al', self.date_to.strftime(DATE_FORMAT)]
        else:
            headersd = ['Al', fields.Date.today().strftime(DATE_FORMAT)]

        for col_num, header in enumerate(headersd, start=start_col):
            sheet.write(0, col_num, header, bold)

        # Encabezados fijos (se han eliminado las columnas: Soporte No. Cheque, Banco, Valor, Fecha Pago)
        headers = [
            'Cuenta', 'Nombre del socio a mostrar', 'Número Factura','Almacén',
            'Fecha de Factura',  'Fecha de Vencimiento', 'Referencia', 'Importe en moneda',
            'En fecha', '1-30', '31-60', '61-90', '91-120', 'Más antiguos'
        ]
        for col_num, header in enumerate(headers):
            sheet.write(2, col_num, header, bold)

        row = self.process_data(account_move_lines, sheet, sheet1, 2, False)
        if self.type == 'advance_account_customer_id':
            report_payment = self.env['account.report'].with_company(company).browse(10)
            # Obtener las opciones
            options_payment = {
                'companies': [{'id': self.company_id.id}],  # Dinámico
                'single_company': [self.company_id.id],
                'date': {
                    'date_from': date_from_formatted,
                    'date_to': date_now_formatted,
                },
                'sections_source_id': 10,
                'export_mode': 'file',
                'all_entries': False,
                'journals': [],
                'show_account': True,
                'comparison': {
                    'filter': 'no_comparison'
                },
                'unfold_all': True  # Propiedad para desplegar todas las cuentas
            }

            # Generar el archivo en binario usando export_to_xlsx para ambas opciones
            xlsx_data = report_payment.export_to_xlsx(options_payment)

            # Leer el archivo binario y cargarlo en openpyxl para la primera opción
            file_content2 = xlsx_data['file_content']
            excel_file2 = io.BytesIO(file_content2)
            workbook2 = openpyxl.load_workbook(excel_file2)

            sheet2 = workbook2.active  # Selecciona la primera hoja

            """Genera y descarga el informe XLSX con la lógica original (pagos, días vencidos, etc.)."""
            self.ensure_one()

            advance_config = self.env['account.advance.config'].search([], limit=1)
            account = advance_config.advance_account_customer_id if advance_config else False
            if account:
                domain_2 = [
                    ('account_id', '=', account.id),
                    ('move_id.state', '=', 'posted'),
                ]
                if self.date_from:
                    domain_2.append(('move_id.date', '>=', self.date_from))
                if self.date_to:
                    domain_2.append(('move_id.date', '<=', self.date_to))
                if self.partner_id:
                    domain_2.append(('partner_id', '=', self.partner_id.id))

                account_move_lines_2 = self.env['account.move.line'].search(domain_2, order='name desc')
                self.process_data(account_move_lines_2, sheet, sheet2, row, True, str(account.code))
            workbook2.close()

        workbook.close()
        workbook1.close()
        output.seek(0)
        xlsx_data = output.read()

        # 4) Convertir a Base64 y asignar a campos binarios
        filename = "informe_pentalab_{}.xlsx".format(fields.Date.today())
        self.file_data = base64.b64encode(xlsx_data)
        self.file_name = filename

        # 5) Retornar acción de descarga
        return {
            'type': 'ir.actions.act_url',
            'url': (
                    '/web/content/%s/%s/file_data?download=true&filename=%s'
                    % (self._name, self.id, self.file_name)
            ),
            'target': 'self',
        }

    def process_data(self, account_move_lines, sheet, sheet1, row, is_payment_account, account_code=''):
        # Ordenar alfabeticamente por el nombre del socio de forma ascendente
        account_move_lines = account_move_lines.sorted(key=lambda aml: aml.partner_id.name or '')
        # 3) Llenar la data
        nombres_sheet1_procesados = set()
        nombres_aml = {aml.partner_id.name: aml for aml in account_move_lines}  # Conjunto de nombres en aml
        ultima_fila_procesada = 4  # Inicializar la última fila procesada
        while ultima_fila_procesada < sheet1.max_row:  # Cambiar a while para controlar el inicio
            row_idx = ultima_fila_procesada + 1  # Ajustar el índice de fila

            nombre_sheet1 = sheet1.cell(row=row_idx, column=1).value  # Nombre principal

            if nombre_sheet1 in nombres_sheet1_procesados:
                continue  # Saltar si el nombre ya fue procesado

            # Encontrar el inicio y fin de las facturas para este nombre
            inicio_facturas = row_idx + 1  # La siguiente fila después del nombre
            fin_facturas = inicio_facturas

            # Detectar el final del bloque de facturas
            while fin_facturas < sheet1.max_row:
                siguiente_fila = sheet1.cell(row=fin_facturas + 1, column=1).value
                if siguiente_fila and "Total" in str(siguiente_fila) and nombre_sheet1 in str(
                        siguiente_fila):
                    fin_facturas += 1
                    break  # Encontramos la fila de total
                else:
                    fin_facturas += 1

            # Obtener el total
            total = sheet1.cell(row=fin_facturas, column=10).value
            exit_while = True
            # Recorrer las filas de facturas
            for factura_idx in range(inicio_facturas, fin_facturas):  # Excluir la fila de total
                valores_factura = [sheet1.cell(row=factura_idx, column=col).value for col in
                                   range(4, 10)]  # Valores
                coincidencia = nombres_aml.get(nombre_sheet1)
                partner_name = nombre_sheet1
                move_name = sheet1.cell(row=factura_idx, column=1).value
                move_date = sheet1.cell(row=factura_idx, column=2).value or ''
                move_account_code = sheet1.cell(row=factura_idx, column=3).value or ''
                move_reference = ''
                move_date_maturity = ''

                if coincidencia:
                    if coincidencia.display_name == sheet1.cell(row=factura_idx, column=1).value:
                        move_name = coincidencia.move_name
                        move_reference = coincidencia.move_id.ref or ''
                        move_account_code = coincidencia.account_id.code
                        move_date_maturity = coincidencia.date_maturity.strftime(
                            DATE_FORMAT) if coincidencia.date_maturity else ''
                if isinstance(total, (int, float)):
                    if total <= 0 or any(isinstance(valor, (int, float)) and valor < 0 for valor in valores_factura):
                        # Verificar valores negativos
                        move_reference = ''
                        move_date_maturity = ''

                name = move_name.split(' ')  # Extrae la primera parte
                if name[0] == 'Fact' or name[0] == 'NotCr':
                    referencia_pago = name[0] + ' ' + name[1]
                else:
                    referencia_pago = name[0]
                pago = self.env['account.move'].search([('name', '=', referencia_pago)], limit=1)

                if pago:
                    move_date = pago.date.strftime(DATE_FORMAT)
                    move_date_maturity = pago.invoice_date_due.strftime(DATE_FORMAT)

                    if referencia_pago.startswith("P"):  # Verifica si referencia_pago comienza con "P"
                        move_date_maturity = move_date
                    if is_payment_account:
                        line = pago.line_ids.filtered(lambda l: move_account_code == account_code)
                        if line:
                            exit_while = False
                            abs_amount_residual = abs(line[0].amount_residual)
                            # Recorremos la lista y buscamos coincidencia
                            for idx, val in enumerate(valores_factura):
                                if isinstance(val, (int, float)) and val != 0:
                                    valores_factura[idx] = val * -1
                        elif not line and factura_idx == fin_facturas - 1:
                            exit_while = True
                            break  # salir del for
                warehouse_name = ''
                if move_name:

                    # Buscar patrones de grupos numéricos separados por guiones
                    matches = re.findall(r'\b\d{3,}\b', move_name)

                    if len(matches) >= 2:
                        entity = matches[0]
                        emission = matches[1]

                        warehouse = self.env['stock.warehouse'].search([
                            ('l10n_ec_entity', '=', entity),
                            ('l10n_ec_emission', '=', emission),
                        ], limit=1)

                        warehouse_name = warehouse.partner_id.city if warehouse else ''
                    else:
                        warehouse_name = ''
                if (is_payment_account and not exit_while) or (not is_payment_account):
                    row = self.write_rows(partner_name, move_name, move_date,
                      move_date_maturity, move_reference, move_account_code, row, sheet,
                      valores_factura, warehouse_name)
            nombres_sheet1_procesados.add(nombre_sheet1)
            ultima_fila_procesada = fin_facturas
            if exit_while:
                continue

        return row

    @staticmethod
    def write_rows(partner_name, move_name, move_date, move_date_maturity, move_reference,
               move_account_code, row, sheet, valores_factura, warehouse_name=''):
        for i, valor in enumerate(valores_factura):
            if isinstance(valor, (int, float)) and valor != 0:
                row += 1  # Incrementar para la nueva fila
                sheet.write(row, 0, move_account_code)
                sheet.write(row, 1, partner_name)
                sheet.write(row, 2, move_name)
                sheet.write(row, 3, warehouse_name)  # nueva columna
                sheet.write(row, 4, move_date)
                sheet.write(row, 5, move_date_maturity)
                sheet.write(row, 6, move_reference)
                sheet.write(row, 7, valor)
                sheet.write(row, 8 + i, valor)
        return row

    def generar_reporte_filtrado_fact(self):
        today = date.today()
        date_from = today.replace(day=1)
        date_to = today

        date_from_str = format_date(self.env, date_from)
        date_to_str = format_date(self.env, date_to)

        options = {
            'companies': [{'id': self.company_id.id}],
            'single_company': [self.company_id.id],
            'report_id': 10,
            'sections_source_id': 10,
            'selected_variant_id': 10,
            'variants_source_id': 10,
            'export_mode': 'file',
            'all_entries': False,
            'journals': [],
            'show_account': True,
            'comparison': {
                'filter': 'no_comparison'
            },
            'unfold_all': True,
            'unreconciled': False,
            'partner': True,
            'fiscal_position': 'all',
            'account_type': [
                {'id': 'trade_payable', 'name': 'Por pagar', 'selected': True},
                {'id': 'non_trade_payable', 'name': 'Cuentas por pagar no comerciales', 'selected': False}
            ],
            'date': {
                'date_from': self.date_from or (self.date_to - timedelta(days=30)) if self.date_to else fields.Date.today() - timedelta(days=30),
                'date_to': self.date_to or fields.Date.today(),
                'filter': 'today',
                'mode': 'single',
                'period_type': 'today',
                'string': f'Al {fields.Date.to_string(self.date_to or fields.Date.today())}',
                'currency_table_period_key': f"None_{fields.Date.to_string(self.date_to or fields.Date.today())}",
            }
        }

        report = self.env['account.report'].with_company(self.company_id).browse(10)

        xlsx_content = report.export_to_xlsx(options)

         # Cargar archivo para edición
        file_content = xlsx_content.get("file_content")
        xlsx_io = BytesIO(file_content)
        wb = load_workbook(filename=xlsx_io)
        ws = wb.active

        # Eliminar filas que no comienzan con 'fact' o 'factura' desde fila 5
        rows_to_delete = []
        for row in range(5, ws.max_row + 1):
            cell_val = str(ws.cell(row=row, column=1).value or '').strip().lower()
            if not (cell_val.startswith("fact") or cell_val.startswith("factura")):
                rows_to_delete.append(row)

        # Eliminar de abajo hacia arriba
        for idx in reversed(rows_to_delete):
            ws.delete_rows(idx)

        new_totals = [0.0] * 7  # columnas D (4) a J (10)

        for row in range(5, ws.max_row + 1):
            for i in range(7):
                val = ws.cell(row=row, column=4 + i).value
                if isinstance(val, (int, float)):
                    new_totals[i] += val

        # Escribir totales en fila 4
        for i, total in enumerate(new_totals):
            ws.cell(row=4, column=4 + i).value = round(total, 2)

        # Escribir suma total en columna J (10), fila 4
        total_general = round(sum(new_totals), 2)
        ws.cell(row=4, column=10).value = total_general

        # Guardar archivo modificado
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        modified_xlsx = output.read()
        
        
        # Guardar como attachment
        attachment = self.env['ir.attachment'].create({
            'name': f'Reporte-Filtrado-{self.company_id.name}-{date_to}.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(modified_xlsx),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Acción para forzar descarga
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
    