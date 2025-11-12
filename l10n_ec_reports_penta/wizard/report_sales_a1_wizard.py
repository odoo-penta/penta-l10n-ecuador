# -*- coding: utf-8 -*-
from odoo import models, fields
import base64
import io
from odoo.tools.misc import xlsxwriter
from odoo.addons.penta_base.reports.xlsx_formats import get_xlsx_formats
from odoo.tools import format_invoice_number
from openpyxl.utils import get_column_letter


class ReportSalesA1Wizard(models.TransientModel):
    _name = 'report.sales.a1.wizard'
    _description = 'Wizard to generate report sales A1'

    def _get_selection_opcions(self):
        options = [('0', 'Todos')]
        types = self.env['l10n_latam.document.type'].search([('active', '=', True)])
        for t in types:
            options.append((str(t.id), t.name))
        return options
    
    date_start = fields.Date(string='Desde', required=True)
    date_end = fields.Date(string='Hasta', required=True)
    document_type = fields.Selection(selection=lambda self: self._get_selection_opcions(), default='0', string='Tipo documento', required=True)
    
    def _get_invoices_data(self):
        # Generar data para reporte
        inv_domain = [
            ('state', '=', 'posted'),
            ('invoice_date', '>=', self.date_start),
            ('invoice_date', '<=', self.date_end),
            ('journal_id.type', '=', 'sale'),
        ]
        if self.document_type != '0':
            inv_domain.append(('l10n_latam_document_type_id', '=', int(self.document_type)))
        invoices = self.env['account.move'].search(inv_domain, order='invoice_date asc')
        return invoices
    
    def print_report(self):
        report = self.generate_xlsx_report()
        today = fields.Date.context_today(self)
        file_name = f"VentasA1_{today.strftime('%d_%m_%Y')}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': file_name,
            'type': 'binary',
            'datas': base64.b64encode(report),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
    
    def generate_xlsx_report(self):
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet("Ventas A1")
        # Formatos
        formats = get_xlsx_formats(workbook)
        # Obtener data
        invoices = self._get_invoices_data()
        # Obtener grupos de impuestos para el reporte
        tax_groups = self.env['account.tax.group'].search([('show_report', '=', True)], order="report_name")
        # Ancho de columnas
        worksheet.set_column('A:A', 6)
        worksheet.set_column('B:C', 24)
        worksheet.set_column('D:D', 15)
        worksheet.set_column('E:E', 30)
        worksheet.set_column('F:G', 20)
        worksheet.set_column('H:I', 22)
        worksheet.set_column('J:J', 15)
        # Ajustar anchos de columna segun cantidad de grupos de impuestos
        last_column = len(tax_groups) * 2 + 10
        worksheet.set_column(get_column_letter(11)+':'+get_column_letter(last_column), 15)
        worksheet.set_column(get_column_letter(last_column+1)+':'+get_column_letter(last_column+2), 15)
        worksheet.set_column(get_column_letter(last_column+3)+':'+get_column_letter(last_column+3), 35)
        worksheet.set_column(get_column_letter(last_column+4)+':'+get_column_letter(last_column+4), 15)
        worksheet.set_column(get_column_letter(last_column+5)+':'+get_column_letter(last_column+5), 37)
        # Encabezados
        headers = ['#', 'TIPO DE COMPROBANTE', 'TIPO DE IDENTIFICACIÓN', 'IDENTIFICACIÓN', 'RAZÓN SOCIAL', 'PARTE RELACIONADA', 'TIPO DE SUJETO', 'NÚMERO DE DOCUMENTO',
                    'NÚMERO AUTORIZACIÓN', 'FECHA EMISIÓN']
        tax_col = 10
        tax_struct = {}
        # Mapear bases
        for tax_group in tax_groups:
            headers.append('BASE ' + tax_group.report_name.upper())
            tax_struct[tax_group.id] = {'base': tax_col}
            tax_col += 1
        # Mapear ivastax_col
        for tax_group in tax_groups:
            headers.append('MONTO ' + tax_group.report_name.upper())
            tax_struct[tax_group.id]['iva'] = tax_col
            tax_col += 1
        # LLenar el resto del texto de la cabecera
        headers += ['TOTAL VENTA', 'RETENCIÓN', 'CASILLA 104', 'DÍAS CRÉDITO', 'FORMA DE PAGO']
        # Mapear cabecera
        company_name = self.env.company.display_name
        worksheet.merge_range('A1:E1', company_name)
        date_from = self.date_start
        worksheet.merge_range('A2:B2', 'Fecha Desde:')
        worksheet.write('C2', date_from.strftime('%d/%m/%Y') if date_from else '')
        date_to = self.date_end
        worksheet.merge_range('A3:B3', 'Fecha Hasta:')
        worksheet.write('C3', date_to.strftime('%d/%m/%Y') if date_to else '')
        worksheet.merge_range('A4:B4', 'Reporte:')
        worksheet.write('C4', 'VENTAS A1')
        row = 5
        # Mapear titulos
        for col, header in enumerate(headers):
            worksheet.write(row, col, header, formats['header_bg'])
        # Mapear datos
        cont = 1
        for invoice in invoices:
            # Obtenemos tags (invoice_line_ids)
            unique_tag_groups = []
            for line in invoice.invoice_line_ids:
                tag_names = sorted(line.tax_tag_ids.mapped('name'))
                if tag_names and tag_names not in unique_tag_groups:
                    unique_tag_groups.append(tag_names)
            # Si no hay tags, ponemos una lista con un solo elemento '' para que haga una fila igual
            tags_to_iterate = [", ".join(tags) for tags in unique_tag_groups] if unique_tag_groups else ['']
            for tag_name in tags_to_iterate:                        
                row += 1
                worksheet.write(row, 0, cont, formats['center'])
                worksheet.write(row, 1, invoice.l10n_latam_document_type_id.name, formats['center'])
                worksheet.write(row, 2, invoice.partner_id.l10n_latam_identification_type_id.name or '', formats['border'])
                worksheet.write(row, 3, invoice.partner_id.vat or '', formats['border'])
                worksheet.write(row, 4, invoice.partner_id.complete_name or '', formats['border'])
                worksheet.write(row, 5, 'SI' if invoice.partner_id.l10n_ec_related_party else 'NO', formats['center'])
                subjet_type = ''
                if invoice.partner_id.company_type == 'person':
                    subjet_type = 'Persona Natural'
                elif invoice.partner_id.company_type == 'company':
                    subjet_type = 'Empresa'
                worksheet.write(row, 6, subjet_type, formats['border'])
                worksheet.write(row, 7, format_invoice_number(invoice.name) or '', formats['border'])
                worksheet.write(row, 8, invoice.l10n_ec_authorization_number or '', formats['border'])
                worksheet.write(row, 9, invoice.invoice_date.strftime("%d/%m/%Y") or '', formats['border'])
                # Mapear impuestos
                base_per_group = {tg.id: 0.0 for tg in tax_groups}
                iva_per_group = {tg.id: 0.0 for tg in tax_groups}
                # Buscar todas las líneas que contienen este mismo tag
                tag_lines = invoice.invoice_line_ids.filtered(
                    lambda l, tag=tag_name: ", ".join(sorted(l.tax_tag_ids.mapped('name'))) == tag
                )
                for line in tag_lines:
                    # Para cada impuesto en la línea
                    for tax in line.tax_ids:
                        tax_group_id = tax.tax_group_id.id
                        if tax_group_id in tax_groups.ids:
                            # Calcular los montos (base e impuesto)
                            taxes_res = tax.compute_all(
                                line.price_unit,
                                currency=invoice.currency_id,
                                quantity=line.quantity,
                                product=line.product_id,
                                partner=invoice.partner_id,
                                rounding_method='round_globally',
                            )
                            base_amount = taxes_res['total_excluded']
                            iva_amount = sum(t['amount'] for t in taxes_res['taxes'])
                            # Sumar al grupo correspondiente
                            base_per_group[tax_group_id] += base_amount
                            iva_per_group[tax_group_id] += iva_amount
                # Escribir bases e impuestos por grupo
                for tg in tax_groups:
                    worksheet.write(row, tax_struct[tg.id]['base'], base_per_group[tg.id], formats['number'])
                    worksheet.write(row, tax_struct[tg.id]['iva'], iva_per_group[tg.id], formats['number'])
                # Total venta
                total_line = sum(base_per_group.values()) + sum(iva_per_group.values())
                worksheet.write(row, tax_col, total_line, formats['number'])    
                # Casilla Retenciones
                has_posted_withhold = any(ret.state == 'posted' for ret in invoice.l10n_ec_withhold_ids)
                worksheet.write(row, tax_col+1, 'SI' if has_posted_withhold else 'NO', formats['center'])
                worksheet.write(row, tax_col+2, tag_name or '', formats['border'])
                worksheet.write(row, tax_col+3, invoice.invoice_payment_term_id.name if invoice.invoice_payment_term_id else '', formats['border'])
                worksheet.write(row, tax_col+4, invoice.l10n_ec_sri_payment_id.name if invoice.l10n_ec_sri_payment_id else '', formats['border'])
                cont += 1
        workbook.close()
        output.seek(0)
        return output.read()